# Файл, содержащий запросы в базу данных


# Подключение модулей Python
from datetime import datetime, date as Date, timezone, timedelta
from dateutil.relativedelta import relativedelta
from aiogram.types import FSInputFile
from sqlalchemy import select, update
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib import colors


# Подключение пользовательских модулей
from bot import bot
from database.models import session, User, Group, Report


# Класс для работы с датой и временем
class Datetime_Handler:
    """Класс для работы с датой и временем"""

    # Статический метод для получения даты в часовом поясе по смещению UTC
    @staticmethod
    async def get_local_date(utc_offset: int) -> Date:
        date: Date = datetime.now(timezone(timedelta(seconds=utc_offset))).date()

        return date

    # Метод класса для получения даты начала текущей недели в часовом поясе по смещению UTC
    @classmethod
    async def get_start_of_week(cls, utc_offset: int) -> Date:
        date: Date = await cls.get_local_date(utc_offset)
        weekday: int = date.weekday()
        date -= relativedelta(days=weekday)

        return date

    # Метод класса для получения даты начала текущего месяца в часовом поясе по смещению UTC
    @classmethod
    async def get_start_of_month(cls, utc_offset: int) -> Date:
        date: Date = (await cls.get_local_date(utc_offset)).replace(day=1)

        return date

    # Метод класса для получения даты начала текущего года в часовом поясе по смещению UTC
    @classmethod
    async def get_start_of_year(cls, utc_offset: int) -> Date:
        date: Date = (await cls.get_local_date(utc_offset)).replace(month=1, day=1)

        return date

    # Статический метод для валидации смещения UTC
    @staticmethod
    async def validate_utc_offset(utc_offset: str) -> bool:
        if not utc_offset[0] in ["+", "-"]:
            return False

        colon_position: int = utc_offset.find(":")

        if colon_position == -1:
            return False

        try:
            hours = int(utc_offset[1:colon_position])

            if not 0 <= hours <= 23:
                return False
        except ValueError:
            return False

        try:
            minutes = int(utc_offset[colon_position + 1 :])

            if not 0 <= minutes <= 59:
                return False
        except ValueError:
            return False

        return True

    # Метод класса для валидации даты в часовом поясе по смещению UTC и самой дате
    @classmethod
    async def validate_date(cls, utc_offset: int, date: str) -> bool:
        try:
            date: Date = datetime.strptime(date, "%d.%m.%Y").date()

            if date > await cls.get_local_date(utc_offset):
                return False

            return True
        except ValueError:
            return False

    # Статический метод для конвертации строкового значения смещения UTC в числовое
    @staticmethod
    async def utc_offset_string_to_int(utc_offset: str) -> int:
        sign: str = utc_offset[0]
        colon_position: str = utc_offset.find(":")
        hours = int(utc_offset[1:colon_position])
        minutes = int(utc_offset[colon_position + 1 :])
        utc_offset_in_seconds: int = hours * 3600 + minutes * 60

        if sign == "-":
            utc_offset_in_seconds = -utc_offset_in_seconds

        return utc_offset_in_seconds


# Класс для записи данных в файл *.xlsx
class Xlsx_Writer:
    """Класс для записи данных в файл *.xlsx"""

    # Статический метод для создания файла отчётов по количеству рядов в нём, данным в рядах и высотам рядов
    @staticmethod
    async def create_reports_file(
        cnt_rows: int,
        rows_data: list[list[str]],
        row_heights: list[int],
    ) -> None:
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = "Отчёты"

        for i in range(cnt_rows):
            work_sheet.append(rows_data[i])

        for i in range(1, cnt_rows + 1):
            work_sheet.row_dimensions[i].height = row_heights[i - 1]

        columns = ["A", "B"]
        for column in columns:
            for cell in work_sheet[column]:
                work_sheet.column_dimensions[column].width = 35

                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                black_color = "000000"
                white_color = "FFFFFF"

                cell.font = Font(name="Times New Roman", size=14, color=black_color)

                cell.fill = PatternFill(
                    start_color=white_color,
                    end_color=white_color,
                    fill_type="solid",
                )

                medium = Side(border_style="medium", color=black_color)
                cell.border = Border(
                    left=medium, right=medium, top=medium, bottom=medium
                )

        work_book.save("./database/Отчёты.xlsx")


# Класс для записи данных в файл *.pdf
class Pdf_Writer:
    """Класс для записи данных в файл *.pdf"""

    # Статический метод для создания файла отчётов по количеству рядов в нём, данным в рядах и высотам рядов
    @staticmethod
    async def create_reports_file(
        cnt_rows: int, rows_data: list[list[str]], rows_heights: dict[int, int]
    ) -> None:
        pdf = SimpleDocTemplate("./database/Отчёты.pdf", pagesize=letter)

        table = Table(rows_data, colWidths=[200, 200, 200], rowHeights=rows_heights)

        pdfmetrics.registerFont(TTFont("Times New Roman", "./database/times.ttf"))
        start = (0, 0)
        end = (-1, -1)
        style = TableStyle(
            [
                ("ALIGN", start, end, "CENTER"),
                ("VALIGN", start, end, "MIDDLE"),
                ("FONTNAME", start, end, "Times New Roman"),
                ("FONTSIZE", start, end, 14),
                ("TEXTCOLOR", start, end, colors.black),
                ("BACKGROUND", start, end, colors.white),
                ("GRID", start, end, 1, colors.black),
            ]
        )

        table.setStyle(style)

        pdf.build([table])


# Класс для описания запросов о пользователе базу данных
class User_Requests:
    """Класс для описания запросов о пользователях в базу данных"""

    # Статический метод для получения объекта пользователя из базы данных по его Телеграм id
    @staticmethod
    async def get(tg_id: int) -> User | None:
        async with session() as sess:
            user: User | None = await sess.scalar(
                select(User).where(User.tg_id == tg_id)
            )

            return user

    # Статический метод для получения Телеграм id пользователя по его id в базе данных
    @staticmethod
    async def get_tg_id(id: int) -> int:
        async with session() as sess:
            user: User | None = await sess.scalar(select(User).where(User.id == id))

            return user.tg_id

    # Метод класса для получения списка групп, которые создал пользователь по его Телеграм id
    @classmethod
    async def get_groups_where_creator(cls, tg_id: int) -> list[str]:
        async with session() as sess:
            id: int = (await cls.get(tg_id)).id

            groups: list[str] = [
                group.name
                for group in await sess.scalars(
                    select(Group).where(Group.creator == id)
                )
            ]
            groups.sort()

            return groups

    # Метод класса для получения списка групп, в которых пользователь назначен получателем отчётов по его Телеграм id
    @classmethod
    async def get_groups_where_reports_recipient(cls, tg_id: int) -> list[str]:
        async with session() as sess:
            id: int = (await cls.get(tg_id)).id

            groups: list[str] = [
                group.name
                for group in await sess.scalars(
                    select(Group).where(Group.reports_recipient == id)
                )
            ]
            groups.sort()

            return groups

    # Метод класса для инициализации пользователя в базе данных по его Телеграм id
    @classmethod
    async def init(cls, tg_id: int) -> None:
        async with session() as sess:
            user: User | None = await cls.get(tg_id)

            if user is None:
                default_utc_offset = 10800

                sess.add(
                    User(tg_id=tg_id, utc_offset=default_utc_offset, feedbacks_cnt=0)
                )

                await sess.commit()

    # Статический метод для установки смещения UTC пользователя в базе данных по его Телеграм id
    @staticmethod
    async def set_utc_offset(tg_id: int, utc_offset: int) -> None:
        async with session() as sess:
            await sess.execute(
                update(User).where(User.tg_id == tg_id).values(utc_offset=utc_offset)
            )

            await sess.commit()

    # Статический метод для увеличения количества отправленных отзывов пользователем
    # в базе данных по его Телеграм id
    @staticmethod
    async def increase_feedbacks_cnt(tg_id: int) -> None:
        async with session() as sess:
            await sess.execute(
                update(User)
                .where(User.tg_id == tg_id)
                .values(feedbacks_cnt=User.feedbacks_cnt + 1)
            )

            await sess.commit()

    # Статический метод для сброса количества отправленных отзывов пользователями в базе данных
    @staticmethod
    async def reset_feedbacks_cnt() -> None:
        async with session() as sess:
            await sess.execute(update(User).values(feedbacks_cnt=0))

            await sess.commit()


# Класс для описания запросов о группах в базу данных
class Group_Requests:
    """Класс для описания запросов о группах в базу данных"""

    # Статический метод для получения объекта группы из базы данных по Телеграм id создателя и её имени
    @staticmethod
    async def get_by_creator(creator_tg_id: int, name: str) -> Group | None:
        async with session() as sess:
            creator: int = (await User_Requests.get(creator_tg_id)).id
            group: Group | None = await sess.scalar(
                select(Group).where(Group.creator == creator).where(Group.name == name)
            )

            return group

    # Статический метод для получения объекта группы из базы данных по Телеграм id получателя отчётов и её имени
    @staticmethod
    async def get_by_reports_recipient(
        name: str, reports_recipient_tg_id: int
    ) -> Group | None:
        async with session() as sess:
            reports_recipient: int = (
                await User_Requests.get(reports_recipient_tg_id)
            ).id
            group: Group | None = await sess.scalar(
                select(Group)
                .where(Group.name == name)
                .where(Group.reports_recipient == reports_recipient)
            )

            return group

    # Статический метод для создания (записи) группы в базе данных по Телеграм id создателя и её имени
    @staticmethod
    async def create(creator_tg_id: int, name: str) -> None:
        async with session() as sess:
            creator: int = (await User_Requests.get(creator_tg_id)).id

            sess.add(
                Group(creator=creator, name=name, reports_recipient=creator, members="")
            )

            await sess.commit()

    # Метод класса для добавления (записи) участника группы в группу в базе данных по Телеграм id создателя группы, её имени и имени участника
    @classmethod
    async def add_member(cls, creator_tg_id: int, name: str, member: str) -> None:
        async with session() as sess:
            creator: int = (await User_Requests.get(creator_tg_id)).id
            members: list[str] = (await cls.get_by_creator(creator_tg_id, name)).members

            if members == "":
                members: list[str] = []
            else:
                members: list[str] = members.split(";\n")

            members.append(member)
            members: str = ";\n".join(members)

            await sess.execute(
                update(Group)
                .where(Group.creator == creator)
                .where(Group.name == name)
                .values(members=members)
            )

            await sess.commit()

    # Метод класса для удаления группы из базы данных по Телеграм id создателя и её имени
    @classmethod
    async def delete(cls, creator_tg_id: int, name: str) -> None:
        async with session() as sess:
            group: Group | None = await cls.get_by_creator(creator_tg_id, name)

            await sess.delete(group)

            await sess.commit()

    # Метод класса для удаления участника из группы в базе данных по Телеграм id создателя группы, её имени и имени участника
    @classmethod
    async def remove_member(cls, creator_tg_id: int, name: str, member: str) -> None:
        async with session() as sess:
            group: Group | None = await cls.get_by_creator(creator_tg_id, name)
            id: int = group.id
            creator: int = (await User_Requests.get(creator_tg_id)).id
            members: list[str] = group.members.split(";\n")
            members.remove(member)
            members: str = ";\n".join(members)

            await sess.execute(
                update(Group)
                .where(Group.creator == creator)
                .where(Group.name == name)
                .values(members=members)
            )

            reports: list[Report] = await sess.scalars(
                select(Report)
                .where(Report.group == id)
                .where(Report.members.like(f"%{member}%"))
            )
            for report in reports:
                report_members: list[str] = report.members.split(";\n")
                report_members.remove(member)
                report_members: str = ";\n".join(report_members)

                await sess.execute(
                    update(Report)
                    .where(Report.id == report.id)
                    .values(members=report_members)
                )

            await sess.commit()

    # Статический метод для назначения получателя отчётов группы в базе данных по Телеграм id создателя, её имени и id получателя
    @staticmethod
    async def assign_reports_recipient(
        creator_tg_id: int, name: str, reports_recipient_tg_id: int
    ) -> None:
        async with session() as sess:
            creator: int = (await User_Requests.get(creator_tg_id)).id
            reports_recipient: int = (
                await User_Requests.get(reports_recipient_tg_id)
            ).id

            await sess.execute(
                update(Group)
                .where(Group.creator == creator)
                .where(Group.name == name)
                .values(reports_recipient=reports_recipient)
            )

            await sess.commit()


# Класс для описания запросов об отчётах группы в базу данных
class Report_Requests:
    """Класс для описания запросов об отчётах группы в базу данных"""

    # Статический метод для получения объекта сегодняшнего отчёта из базы данных по Телеграм id создателя группы и её имени
    @staticmethod
    async def get(group_creator_tg_id: int, group_name: str) -> Report:
        async with session() as sess:
            user: User | None = await User_Requests.get(group_creator_tg_id)
            group: int = (
                await Group_Requests.get_by_creator(group_creator_tg_id, group_name)
            ).id
            date: Date = await Datetime_Handler.get_local_date(user.utc_offset)
            report: Report | None = await sess.scalar(
                select(Report).where(Report.group == group).where(Report.date == date)
            )

            return report

    # Статический метод для получения статистики об отсутствии участников группы по её имени, Телеграм id получателя отчётов и
    # датам начала и конца периода времени, для которого она получается
    @staticmethod
    async def get_statistics(
        group_name: str,
        reports_recipient_tg_id: int,
        date_from: str,
        date_to: str = None,
    ) -> str:
        async with session() as sess:
            group: Group = await Group_Requests.get_by_reports_recipient(
                group_name, reports_recipient_tg_id
            )
            group_id: int = group.id
            reports_recipient_utc_offset: int = (
                await User_Requests.get(reports_recipient_tg_id)
            ).utc_offset

            if date_from == "Неделя":
                date_from: Date = await Datetime_Handler.get_start_of_week(
                    reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    reports_recipient_utc_offset
                )
            elif date_from == "Месяц":
                date_from: Date = await Datetime_Handler.get_start_of_month(
                    reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    reports_recipient_utc_offset
                )
            elif date_from == "Год":
                date_from: Date = await Datetime_Handler.get_start_of_year(
                    reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    reports_recipient_utc_offset
                )
            else:
                date_from: Date = datetime.strptime(date_from, "%d.%m.%Y").date()
                date_to: Date = datetime.strptime(date_to, "%d.%m.%Y").date()

            reports = await sess.scalars(
                select(Report)
                .where(Report.group == group_id)
                .where(Report.date >= date_from)
                .where(Report.date <= date_to)
            )
            reports_cnt: int = 0
            reports_with_member: dict[str, int] = {}

            for report in reports:
                reports_cnt += 1
                report_members: list[str] = report.members.split(";\n")

                if report_members == [""]:
                    continue

                for report_member in report_members:
                    if not report_member in reports_with_member:
                        reports_with_member[report_member] = 0

                    reports_with_member[report_member] += 1

            if reports_cnt == 0:
                statistics = f'С {date_from.strftime("%d.%m.%Y")} по {date_to.strftime("%d.%m.%Y")} в группе "{group_name}" не создавалось отчётов об отсутствии.'

                return statistics

            reports_with_member_sorted: list = sorted(
                reports_with_member.items(),
                key=lambda item: item[1],
                reverse=True,
            )
            reports_members_cnt: int = len(reports_with_member_sorted)

            if reports_members_cnt == 0:
                statistics = f'С {date_from.strftime("%d.%m.%Y")} по {date_to.strftime("%d.%m.%Y")} в группе "{group_name}" отсутствующих не было.'

                return statistics
            else:
                statistics = f'Статистика отсутствия участников группы "{group_name}" с {date_from.strftime("%d.%m.%Y")} по {date_to.strftime("%d.%m.%Y")}:\n'

                for i in range(reports_members_cnt):
                    member: str = reports_with_member_sorted[i][0]
                    reports_with_this_member: int = reports_with_member_sorted[i][1]
                    reports_with_this_member_percentages = int(
                        reports_with_this_member / reports_cnt * 100
                    )

                    statistics += f"{i + 1}. {member} - Присутствовал в {reports_with_this_member} отчётах из {reports_cnt} "
                    statistics += f"({reports_with_this_member_percentages}%)"

                    if i < reports_members_cnt - 1:
                        statistics += ";\n"
                    else:
                        statistics += "."

                return statistics

    # Статический метод для получения файла отчётов из базы данных по Telegram id создателя группы, её имени и датам начала и конца промежутка времени,
    # для которого получается файл
    @staticmethod
    async def get_file(
        group_name: str,
        group_reports_recipient_tg_id: int,
        date_from: str,
        date_to: str,
        file_format: str,
    ) -> tuple[FSInputFile | str, str]:
        async with session() as sess:
            group: Group = await Group_Requests.get_by_reports_recipient(
                group_name, group_reports_recipient_tg_id
            )
            group_id: int = group.id
            group_creator_name: str = (
                await bot.get_chat(await User_Requests.get_tg_id(group.creator))
            ).username
            group_reports_recipient_utc_offset: int = (
                await User_Requests.get(group_reports_recipient_tg_id)
            ).utc_offset

            if date_from == "Неделя":
                date_from: Date = await Datetime_Handler.get_start_of_week(
                    group_reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    group_reports_recipient_utc_offset
                )
            elif date_from == "Месяц":
                date_from: Date = await Datetime_Handler.get_start_of_month(
                    group_reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    group_reports_recipient_utc_offset
                )
            elif date_from == "Год":
                date_from: Date = await Datetime_Handler.get_start_of_year(
                    group_reports_recipient_utc_offset
                )
                date_to: Date = await Datetime_Handler.get_local_date(
                    group_reports_recipient_utc_offset
                )
            else:
                date_from: Date = datetime.strptime(date_from, "%d.%m.%Y").date()
                date_to: Date = datetime.strptime(date_to, "%d.%m.%Y").date()

            rows_data: list[list[str]] = [
                [
                    "Имя создателя группы",
                    "Название группы",
                ],
                [
                    group_creator_name,
                    group_name,
                ],
                [
                    "Дата создания отчёта",
                    "Участники отчёта",
                ],
            ]
            rows_heights: list[int] = [25, 25, 25]

            reports = await sess.scalars(
                select(Report)
                .where(Report.group == group_id)
                .where(Report.date >= date_from)
                .where(Report.date <= date_to)
                .order_by(Report.date)
            )

            reports_cnt = 0

            i = 4

            for report in reports:
                reports_cnt += 1
                report_date: Date = report.date.strftime("%d.%m.%Y")
                report_members: str = report.members.replace(";\n", "\n")
                lines_cnt = report_members.count("\n") + 1

                rows_data.append([report_date, report_members])
                rows_heights.append(lines_cnt * 25)

                i += 1

            if reports_cnt == 0:
                mssg_txt = f'С {date_from.strftime("%d.%m.%Y")} по {date_to.strftime("%d.%m.%Y")} в группе "{group_name}" не создавалось отчётов об отсутствии.'

                return mssg_txt, ""

            if file_format == "Xlsx":
                await Xlsx_Writer.create_reports_file(
                    reports_cnt + 3, rows_data, rows_heights
                )
            else:
                await Pdf_Writer.create_reports_file(
                    reports_cnt + 3, rows_data, rows_heights
                )

            file = FSInputFile(f"./database/Отчёты.{file_format.lower()}")
            mssg_txt = f'Файл отчётов об отсутствии участников группы "{group_name}" с {date_from.strftime("%d.%m.%Y")} по {date_to.strftime("%d.%m.%Y")}.'

            return file, mssg_txt

    # Статический метод для создания сегодняшнего отчёта в базе данных по Телеграм id создателя группы, её имени и списку участников отчёта
    async def create(
        group_creator_tg_id: int, group_name: str, members: list[str]
    ) -> None:
        async with session() as sess:
            user: User | None = await User_Requests.get(group_creator_tg_id)
            group: int = (
                await Group_Requests.get_by_creator(group_creator_tg_id, group_name)
            ).id
            date: Date = await Datetime_Handler.get_local_date(user.utc_offset)
            members: str = ";\n".join(members)

            sess.add(
                Report(
                    group=group,
                    date=date,
                    members=members,
                )
            )

            await sess.commit()

    # Статический метод для редактирования отчёта в базе данных по Телеграм id создателя группы, её имени, дате создания отчёта и списку участников отчёта
    async def edit(
        group_creator_tg_id: int, group_name: str, members: list[str]
    ) -> None:
        async with session() as sess:
            user: User | None = await User_Requests.get(group_creator_tg_id)
            group: int = (
                await Group_Requests.get_by_creator(group_creator_tg_id, group_name)
            ).id
            date: Date = await Datetime_Handler.get_local_date(user.utc_offset)
            members: str = ";\n".join(members)

            await sess.execute(
                update(Report)
                .where(Report.group == group)
                .where(Report.date == date)
                .values(members=members)
            )

            await sess.commit()
